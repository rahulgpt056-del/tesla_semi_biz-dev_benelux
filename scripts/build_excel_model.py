"""
Generate the Tesla Semi Benelux TCO Excel model (live formulas + named ranges).
Mirrors engine/tco_engine.py so the spreadsheet and the web app agree.

Output: model/Tesla-Semi-Benelux-TCO-Model.xlsx
Run:    python scripts/build_excel_model.py
"""
import json
import os

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA = os.path.join(ROOT, "data", "assumptions.json")
OUT_DIR = os.path.join(ROOT, "model")
OUT = os.path.join(OUT_DIR, "Tesla-Semi-Benelux-TCO-Model.xlsx")

a = json.load(open(DATA, encoding="utf-8"))
ep, pol, veh = a["energy_prices"], a["policy"], a["vehicle"]

NAVY = "FF0B2545"
RED = "FFE31937"
LIGHT = "FFEFF3F8"
INPUTBG = "FFFFF3D6"
WHITE = "FFFFFFFF"
GREY = "FF5B6670"

thin = Side(style="thin", color="FFD2DAE2")
box = Border(left=thin, right=thin, top=thin, bottom=thin)
title_font = Font(name="Arial", size=16, bold=True, color=NAVY)
hdr_font = Font(name="Arial", size=11, bold=True, color=WHITE)
lbl_font = Font(name="Arial", size=10, color="FF1A1A1A")
val_font = Font(name="Arial", size=10, bold=True, color=NAVY)
note_font = Font(name="Arial", size=9, italic=True, color=GREY)
big_font = Font(name="Arial", size=20, bold=True, color=NAVY)

wb = Workbook()

# --------------------------------------------------------------------------- #
# INPUTS sheet
# --------------------------------------------------------------------------- #
ws = wb.active
ws.title = "Inputs"
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 38
ws.column_dimensions["B"].width = 16
ws.column_dimensions["C"].width = 52

ws["A1"] = "Tesla Semi · Benelux — TCO Model — Inputs"
ws["A1"].font = title_font
ws["A2"] = "Edit the amber cells. Everything recalculates on the Model and Dashboard sheets."
ws["A2"].font = note_font

names: dict[str, str] = {}
row = 4


def section(title: str):
    global row
    ws.cell(row=row, column=1, value=title)
    for c in range(1, 4):
        cell = ws.cell(row=row, column=c)
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.font = hdr_font
    row += 1


def inp(label, value, name, note="", number_format="0.###"):
    global row
    ws.cell(row=row, column=1, value=label).font = lbl_font
    vcell = ws.cell(row=row, column=2, value=value)
    vcell.font = val_font
    vcell.fill = PatternFill("solid", fgColor=INPUTBG)
    vcell.border = box
    vcell.number_format = number_format
    vcell.alignment = Alignment(horizontal="right")
    if note:
        ws.cell(row=row, column=3, value=note).font = note_font
    names[name] = f"Inputs!$B${row}"
    row += 1


section("Fleet & duty cycle")
inp("Fleet size (trucks)", 10, "FleetSize", number_format="0")
inp("Daily distance per truck (km)", 350, "DailyKm", number_format="0")
inp("Operating days / year", 300, "OpDays", number_format="0")
inp("Average payload (t)", 22, "Payload", number_format="0")
inp("Route factor (urban 1.15 / regional 1.00 / long-haul 1.08)", 1.00, "RouteFactor",
    "Multiplier on base kWh/km for the route profile.", "0.00")
inp("Charging done at depot (fraction)", 1.00, "DepotShare", "1.00 = all charging at depot.", "0.00")
inp("Depot dwell / charging window (h)", 9, "Dwell", number_format="0.0")

section("Energy & prices")
inp("Semi consumption (kWh/km, base)", veh["semi_kwh_per_km_base"], "SemiKwh", number_format="0.00")
inp("Electricity price (EUR/kWh)", ep["electricity_eur_per_kwh"], "ElecPrice", number_format="0.000")
inp("Diesel price (EUR/L)", ep["diesel_eur_per_litre"], "DieselPrice", number_format="0.00")
inp("Diesel use (L/100km)", veh["diesel_l_per_100km"], "DieselL100", number_format="0.0")

section("Charging & site")
inp("Charger power (kW)", 350, "ChargerKw", "Depot charger; MCS Megacharger up to 1200 kW.", "0")
inp("Depot grid capacity (kW)", 1400, "DepotKw", number_format="0")
inp("Semi single-charge range (km)", veh["semi_range_km"], "SemiRange", number_format="0")

section("Capex, opex & incentives")
inp("Semi acquisition cost (EUR/truck)", 230000, "SemiCapex", number_format="#,##0")
inp("Diesel equivalent (EUR/truck)", 130000, "DieselCapex", number_format="#,##0")
inp("Semi maintenance (EUR/km)", 0.12, "SemiMaintKm", number_format="0.00")
inp("Diesel maintenance (EUR/km)", 0.19, "DieselMaintKm", number_format="0.00")
inp("AanZET subsidy (EUR/truck)", pol["aanzet_subsidy_eur_per_truck"], "Aanzet", number_format="#,##0")
inp("Truck toll — vrachtwagenheffing (EUR/km)", pol["truck_toll_eur_per_km"], "TollKm", number_format="0.000")
inp("Zero-emission also pays toll? (1=yes,0=no)", 0, "ZEPaysToll", number_format="0")

section("Analysis & constants")
inp("Analysis horizon (years)", 5, "Years", number_format="0")
inp("Payload reference (t)", 19, "PayloadRef", "Energy rises above this payload.", "0")
inp("Payload sensitivity (per t)", 0.006, "PayloadSens", number_format="0.000")
inp("Charge efficiency", 0.92, "ChargeEff", number_format="0.00")
inp("Usable grid capacity fraction", 0.90, "UsableCap", number_format="0.00")

# register defined names
for nm, ref in names.items():
    wb.defined_names.add(DefinedName(nm, attr_text=ref))

# --------------------------------------------------------------------------- #
# MODEL sheet (formulas)
# --------------------------------------------------------------------------- #
ms = wb.create_sheet("Model")
ms.sheet_view.showGridLines = False
ms.column_dimensions["A"].width = 42
ms.column_dimensions["B"].width = 20
ms.column_dimensions["C"].width = 10
ms["A1"] = "Model — calculated outputs (formulas reference the Inputs named ranges)"
ms["A1"].font = title_font

mrow = 3
mnames: dict[str, str] = {}


def calc(label, formula, name, number_format="#,##0.##", unit=""):
    global mrow
    ms.cell(row=mrow, column=1, value=label).font = lbl_font
    c = ms.cell(row=mrow, column=2, value=formula)
    c.font = val_font
    c.number_format = number_format
    c.border = box
    c.fill = PatternFill("solid", fgColor=LIGHT)
    if unit:
        ms.cell(row=mrow, column=3, value=unit).font = note_font
    mnames[name] = f"Model!$B${mrow}"
    mrow += 1


calc("Effective consumption", "=SemiKwh*RouteFactor*(1+MAX(0,Payload-PayloadRef)*PayloadSens)",
     "EffKwh", "0.000", "kWh/km")
calc("Daily energy per truck", "=DailyKm*EffKwh", "DailyPerTruck", "#,##0", "kWh")
calc("Fleet energy per day", "=DailyPerTruck*FleetSize*DepotShare", "FleetDaily", "#,##0", "kWh")
calc("Fleet energy per year", "=FleetDaily*OpDays", "AnnualEnergy", "#,##0", "kWh")
calc("Fleet km per year", "=DailyKm*OpDays*FleetSize", "FleetAnnualKm", "#,##0", "km")
calc("Diesel litres displaced / yr", "=FleetAnnualKm*DieselL100/100", "DieselLitresYr", "#,##0", "L")
calc("Charge time per truck", "=(DailyPerTruck*DepotShare)/(ChargerKw*ChargeEff)", "ChargeTime", "0.00", "h")
calc("Trucks per charger (in window)", "=MAX(1,Dwell/ChargeTime)", "TrucksPerCharger", "0.0", "")
calc("Chargers needed", "=MAX(1,CEILING(FleetSize/TrucksPerCharger,1))", "ChargersNeeded", "0", "")
calc("Peak charging load", "=ChargersNeeded*ChargerKw", "PeakLoad", "#,##0", "kW")
calc("Usable depot capacity", "=DepotKw*UsableCap", "UsableCapacity", "#,##0", "kW")
calc("Grid capacity status", '=IF(PeakLoad<=UsableCapacity,"OK","Upgrade needed")', "CapacityOK", "@", "")
calc("Duty-cycle fit",
     '=IF(DailyKm>SemiRange,"No fit",IF(ChargeTime>Dwell,"Conditional",'
     'IF(PeakLoad>UsableCapacity,"Conditional",IF(DailyKm>0.85*SemiRange,"Conditional","Fit"))))',
     "DutyFit", "@", "")

ms.cell(row=mrow, column=1, value="— Economics —").font = Font(name="Arial", bold=True, color=NAVY)
mrow += 1
calc("Annual electricity cost", "=AnnualEnergy*ElecPrice", "AnnualEnergyCost", "#,##0", "EUR")
calc("Annual diesel fuel cost", "=DieselLitresYr*DieselPrice", "AnnualDieselFuel", "#,##0", "EUR")
calc("Semi toll / yr", "=IF(ZEPaysToll=1,TollKm*FleetAnnualKm,0)", "SemiToll", "#,##0", "EUR")
calc("Diesel toll / yr", "=TollKm*FleetAnnualKm", "DieselToll", "#,##0", "EUR")
calc("Semi net capex (fleet)", "=(SemiCapex-Aanzet)*FleetSize", "SemiCapexNet", "#,##0", "EUR")
calc("Diesel capex (fleet)", "=DieselCapex*FleetSize", "DieselCapexTotal", "#,##0", "EUR")
calc("Semi opex / yr", "=AnnualEnergyCost+SemiMaintKm*FleetAnnualKm+SemiToll", "SemiOpexYr", "#,##0", "EUR")
calc("Diesel opex / yr", "=AnnualDieselFuel+DieselMaintKm*FleetAnnualKm+DieselToll", "DieselOpexYr", "#,##0", "EUR")
calc("Semi TCO (horizon)", "=SemiCapexNet+SemiOpexYr*Years", "SemiTCO", "#,##0", "EUR")
calc("Diesel TCO (horizon)", "=DieselCapexTotal+DieselOpexYr*Years", "DieselTCO", "#,##0", "EUR")
calc("TCO saving (Diesel − Semi)", "=DieselTCO-SemiTCO", "Saving", "#,##0", "EUR")
calc("Semi cost per km", "=SemiTCO/(FleetAnnualKm*Years)", "SemiCPK", "0.000", "EUR/km")
calc("Diesel cost per km", "=DieselTCO/(FleetAnnualKm*Years)", "DieselCPK", "0.000", "EUR/km")
calc("Annual opex saving", "=DieselOpexYr-SemiOpexYr", "AnnualOpexSaving", "#,##0", "EUR")
calc("Net incremental capex", "=SemiCapexNet-DieselCapexTotal", "NetIncrCapex", "#,##0", "EUR")
calc("Simple payback",
     '=IF(NetIncrCapex<=0,0,IF(AnnualOpexSaving<=0,"No payback",NetIncrCapex/AnnualOpexSaving))',
     "SimplePayback", "0.0", "years")

for nm, ref in mnames.items():
    wb.defined_names.add(DefinedName(nm, attr_text=ref))

# --------------------------------------------------------------------------- #
# DASHBOARD sheet
# --------------------------------------------------------------------------- #
ds = wb.create_sheet("Dashboard")
ds.sheet_view.showGridLines = False
for col, w in {"A": 34, "B": 22, "C": 4, "D": 18, "E": 18}.items():
    ds.column_dimensions[col].width = w
ds["A1"] = "Tesla Semi · Benelux — Business Case Dashboard"
ds["A1"].font = title_font
ds["A2"] = "Headline outputs. All figures flow from the Inputs sheet."
ds["A2"].font = note_font


def kpi(r, label, formula, fmt="#,##0"):
    ds.cell(row=r, column=1, value=label).font = lbl_font
    c = ds.cell(row=r, column=2, value=formula)
    c.font = big_font
    c.number_format = fmt
    c.alignment = Alignment(horizontal="right")


kpi(4, "Duty-cycle fit", "=DutyFit", "@")
kpi(5, f"TCO saving over horizon (EUR)", "=Saving")
kpi(6, "Simple payback (years)", "=SimplePayback", "0.0")
kpi(7, "Fleet energy / year (kWh)", "=AnnualEnergy")
kpi(8, "Chargers needed", "=ChargersNeeded", "0")
kpi(9, "Peak load vs capacity", '=PeakLoad&" / "&UsableCapacity&" kW ("&CapacityOK&")"', "@")
kpi(10, "Semi €/km vs Diesel €/km", '=TEXT(SemiCPK,"0.000")&" vs "&TEXT(DieselCPK,"0.000")', "@")

# small comparison table for the chart
ds["D4"] = "Powertrain"; ds["E4"] = "TCO (EUR)"
ds["D4"].font = ds["E4"].font = Font(name="Arial", bold=True, color=NAVY)
ds["D5"] = "Tesla Semi"; ds["E5"] = "=SemiTCO"
ds["D6"] = "Diesel"; ds["E6"] = "=DieselTCO"
for cell in ("E5", "E6"):
    ds[cell].number_format = "#,##0"

chart = BarChart()
chart.type = "col"
chart.title = "5-year TCO: Semi vs Diesel"
chart.legend = None
chart.height = 7
chart.width = 12
data = Reference(ds, min_col=5, min_row=4, max_row=6)
cats = Reference(ds, min_col=4, min_row=5, max_row=6)
chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)
ds.add_chart(chart, "A12")

# --------------------------------------------------------------------------- #
# SOURCES sheet
# --------------------------------------------------------------------------- #
src = wb.create_sheet("Sources")
src.sheet_view.showGridLines = False
src.column_dimensions["A"].width = 40
src.column_dimensions["B"].width = 80
src["A1"] = "Sources & assumptions"
src["A1"].font = title_font
rows = [
    ("Semi efficiency / range / Megacharger", veh.get("semi_efficiency_notes", "")),
    ("AanZET subsidy", pol.get("aanzet_notes", "")),
    ("Truck toll (vrachtwagenheffing)", pol.get("truck_toll_notes", "")),
    ("Electricity & diesel prices", a["energy_prices"].get("_sources", "")),
    ("Last updated", a.get("_meta", {}).get("last_updated", "")),
    ("Note", "Public, illustrative defaults for scoping. Confirm against primary sources before contracting."),
]
for i, (k, v) in enumerate(rows, start=3):
    src.cell(row=i, column=1, value=k).font = val_font
    cc = src.cell(row=i, column=2, value=v)
    cc.font = note_font
    cc.alignment = Alignment(wrap_text=True, vertical="top")

os.makedirs(OUT_DIR, exist_ok=True)
wb.save(OUT)
print("WROTE", OUT)
